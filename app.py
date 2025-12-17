# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, send_file, Response
from datetime import datetime, date
import sqlite3
import json
import os

app = Flask(__name__)
DB_PATH = 'data/alquileres.db'

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    os.makedirs('data', exist_ok=True)
    conn = get_db()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS propiedades (
        id INTEGER PRIMARY KEY,
        nombre TEXT UNIQUE,
        tipo TEXT,
        activo INTEGER DEFAULT 1
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS ocupaciones (
        id INTEGER PRIMARY KEY,
        propiedad_id INTEGER,
        fecha DATE,
        precio REAL,
        origen TEXT,
        notas TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (propiedad_id) REFERENCES propiedades(id),
        UNIQUE(propiedad_id, fecha)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS gastos (
        id INTEGER PRIMARY KEY,
        propiedad_id INTEGER,
        fecha DATE,
        monto REAL,
        categoria TEXT,
        descripcion TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (propiedad_id) REFERENCES propiedades(id)
    )''')
    
    # Tabla para alquileres mensuales (Brickell, locales)
    c.execute('''CREATE TABLE IF NOT EXISTS alquileres_mensuales (
        id INTEGER PRIMARY KEY,
        propiedad_id INTEGER,
        año INTEGER,
        mes INTEGER,
        monto REAL,
        notas TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (propiedad_id) REFERENCES propiedades(id),
        UNIQUE(propiedad_id, año, mes)
    )''')
    
    propiedades = [
        ('TIDES 14 B', 'temporario'),
        ('TIDES 5 L', 'temporario'),
        ('TIDES 10 L', 'temporario'),
        ('TIDES 10 F', 'temporario'),
        ('TIDES 12 F', 'temporario'),
        ('Brickell', 'mensual'),
        ('Local 1', 'mensual'),
        ('Local 2', 'mensual'),
    ]
    for nombre, tipo in propiedades:
        c.execute('INSERT OR IGNORE INTO propiedades (nombre, tipo) VALUES (?, ?)', (nombre, tipo))
    
    # Actualizar tipos existentes
    c.execute("UPDATE propiedades SET tipo = 'temporario' WHERE nombre LIKE 'TIDES%'")
    c.execute("UPDATE propiedades SET tipo = 'mensual' WHERE nombre IN ('Brickell', 'Local 1', 'Local 2')")
    
    conn.commit()
    conn.close()

init_db()

@app.route('/')
def index():
    return render_template('index.html')

# === FORMULARIOS EXTERNOS PARA ALICIA Y ESTANISLAO ===

@app.route('/cargar/<nombre>')
def formulario_externo(nombre):
    if nombre.lower() not in ['alicia', 'estanislao']:
        return "Acceso no autorizado", 403
    return render_template('cargar_externo.html', nombre=nombre.capitalize())

@app.route('/api/cargar-externo', methods=['POST'])
def guardar_carga_externa():
    data = request.json
    conn = get_db()
    
    try:
        # Obtener ID de propiedad
        prop = conn.execute('SELECT id FROM propiedades WHERE nombre = ?', (data['propiedad'],)).fetchone()
        if not prop:
            return jsonify({'success': False, 'error': 'Propiedad no encontrada'}), 400
        
        # Guardar cada fecha del rango
        from datetime import datetime, timedelta
        
        fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d')
        fecha_fin = datetime.strptime(data['fecha_fin'], '%Y-%m-%d')
        
        dias_guardados = 0
        fecha_actual = fecha_inicio
        
        while fecha_actual <= fecha_fin:
            fecha_str = fecha_actual.strftime('%Y-%m-%d')
            
            # Verificar si ya existe una ocupación en esa fecha para esa propiedad
            existente = conn.execute('''
                SELECT id, origen FROM ocupaciones 
                WHERE propiedad_id = ? AND fecha = ?
            ''', (prop['id'], fecha_str)).fetchone()
            
            if existente:
                # Ya existe, no sobrescribir
                fecha_actual += timedelta(days=1)
                continue
            
            # Insertar nueva ocupación
            conn.execute('''
                INSERT INTO ocupaciones (propiedad_id, fecha, precio, origen, notas)
                VALUES (?, ?, ?, ?, ?)
            ''', (prop['id'], fecha_str, data['precio'], data['origen'], data['inquilino']))
            
            dias_guardados += 1
            fecha_actual += timedelta(days=1)
        
        conn.commit()
        return jsonify({'success': True, 'dias': dias_guardados})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/mis-cargas/<origen>')
def obtener_cargas_externo(origen):
    conn = get_db()
    cargas = conn.execute('''
        SELECT o.id, o.fecha, o.precio, o.notas, p.nombre as propiedad
        FROM ocupaciones o
        JOIN propiedades p ON o.propiedad_id = p.id
        WHERE o.origen = ?
        ORDER BY o.fecha DESC
        LIMIT 50
    ''', (origen.capitalize(),)).fetchall()
    conn.close()
    return jsonify([dict(c) for c in cargas])

@app.route('/api/borrar-carga/<int:id>/<origen>', methods=['DELETE'])
def borrar_carga_externa(id, origen):
    conn = get_db()
    # Solo permitir borrar si el origen coincide
    carga = conn.execute('SELECT origen FROM ocupaciones WHERE id = ?', (id,)).fetchone()
    if not carga or carga['origen'].lower() != origen.lower():
        conn.close()
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    conn.execute('DELETE FROM ocupaciones WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/modificar-carga/<int:id>', methods=['PUT'])
def modificar_carga_externa(id):
    data = request.json
    conn = get_db()
    
    # Verificar que el origen coincida
    carga = conn.execute('SELECT origen FROM ocupaciones WHERE id = ?', (id,)).fetchone()
    if not carga or carga['origen'].lower() != data['origen'].lower():
        conn.close()
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    conn.execute('''
        UPDATE ocupaciones SET precio = ?, notas = ? WHERE id = ?
    ''', (data['precio'], data['inquilino'], id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/propiedades')
def get_propiedades():
    conn = get_db()
    props = conn.execute('SELECT * FROM propiedades WHERE activo = 1').fetchall()
    conn.close()
    return jsonify([dict(p) for p in props])

@app.route('/api/ocupaciones/<int:year>/<int:month>')
def get_ocupaciones(year, month):
    conn = get_db()
    ocupaciones = conn.execute('''
        SELECT o.*, p.nombre as propiedad_nombre 
        FROM ocupaciones o 
        JOIN propiedades p ON o.propiedad_id = p.id
        WHERE strftime('%Y', o.fecha) = ? AND strftime('%m', o.fecha) = ?
    ''', (str(year), str(month).zfill(2))).fetchall()
    conn.close()
    return jsonify([dict(o) for o in ocupaciones])

@app.route('/api/ocupacion', methods=['POST'])
def guardar_ocupacion():
    data = request.json
    conn = get_db()
    try:
        conn.execute('''
            INSERT OR REPLACE INTO ocupaciones (propiedad_id, fecha, precio, origen, notas)
            VALUES (?, ?, ?, ?, ?)
        ''', (data['propiedad_id'], data['fecha'], data['precio'], data['origen'], data.get('notas', '')))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/ocupacion/<int:propiedad_id>/<fecha>', methods=['DELETE'])
def eliminar_ocupacion(propiedad_id, fecha):
    conn = get_db()
    conn.execute('DELETE FROM ocupaciones WHERE propiedad_id = ? AND fecha = ?', (propiedad_id, fecha))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/gastos', methods=['GET', 'POST'])
def gastos():
    conn = get_db()
    if request.method == 'POST':
        data = request.json
        conn.execute('''
            INSERT INTO gastos (propiedad_id, fecha, monto, categoria, descripcion)
            VALUES (?, ?, ?, ?, ?)
        ''', (data.get('propiedad_id'), data['fecha'], data['monto'], data['categoria'], data.get('descripcion', '')))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    else:
        year = request.args.get('year', datetime.now().year)
        gastos = conn.execute('''
            SELECT g.*, p.nombre as propiedad_nombre 
            FROM gastos g 
            LEFT JOIN propiedades p ON g.propiedad_id = p.id
            WHERE strftime('%Y', g.fecha) = ?
            ORDER BY g.fecha DESC
        ''', (str(year),)).fetchall()
        conn.close()
        return jsonify([dict(g) for g in gastos])

@app.route('/api/gasto/<int:id>', methods=['DELETE'])
def eliminar_gasto(id):
    conn = get_db()
    conn.execute('DELETE FROM gastos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# === ALQUILERES MENSUALES (Brickell, Local 1, Local 2) ===

@app.route('/api/alquileres-mensuales/<int:year>')
def get_alquileres_mensuales(year):
    conn = get_db()
    alquileres = conn.execute('''
        SELECT a.*, p.nombre as propiedad_nombre 
        FROM alquileres_mensuales a 
        JOIN propiedades p ON a.propiedad_id = p.id
        WHERE a.año = ?
        ORDER BY a.mes
    ''', (year,)).fetchall()
    conn.close()
    return jsonify([dict(a) for a in alquileres])

@app.route('/api/alquiler-mensual', methods=['POST'])
def guardar_alquiler_mensual():
    data = request.json
    conn = get_db()
    try:
        conn.execute('''
            INSERT OR REPLACE INTO alquileres_mensuales (propiedad_id, año, mes, monto, notas)
            VALUES (?, ?, ?, ?, ?)
        ''', (data['propiedad_id'], data['año'], data['mes'], data['monto'], data.get('notas', '')))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/alquiler-mensual/<int:propiedad_id>/<int:anio>/<int:mes>', methods=['DELETE'])
def eliminar_alquiler_mensual(propiedad_id, anio, mes):
    conn = get_db()
    conn.execute('DELETE FROM alquileres_mensuales WHERE propiedad_id = ? AND año = ? AND mes = ?', 
                 (propiedad_id, anio, mes))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/resumen/<int:year>')
def resumen(year):
    conn = get_db()
    
    # Ingresos de ocupaciones temporarias
    ingresos = conn.execute('''
        SELECT p.id, p.nombre, p.tipo, o.origen,
               COUNT(*) as noches,
               SUM(o.precio) as total_ingresos
        FROM ocupaciones o
        JOIN propiedades p ON o.propiedad_id = p.id
        WHERE strftime('%Y', o.fecha) = ?
        GROUP BY p.id, o.origen
    ''', (str(year),)).fetchall()
    
    # Ingresos de alquileres mensuales
    ingresos_mensuales = conn.execute('''
        SELECT p.id, p.nombre, p.tipo,
               COUNT(*) as meses,
               SUM(a.monto) as total_ingresos
        FROM alquileres_mensuales a
        JOIN propiedades p ON a.propiedad_id = p.id
        WHERE a.año = ?
        GROUP BY p.id
    ''', (year,)).fetchall()
    
    gastos = conn.execute('''
        SELECT p.id, p.nombre, g.categoria,
               SUM(g.monto) as total_gastos
        FROM gastos g
        LEFT JOIN propiedades p ON g.propiedad_id = p.id
        WHERE strftime('%Y', g.fecha) = ?
        GROUP BY p.id, g.categoria
    ''', (str(year),)).fetchall()
    
    gastos_generales = conn.execute('''
        SELECT SUM(monto) as total
        FROM gastos
        WHERE propiedad_id IS NULL AND strftime('%Y', fecha) = ?
    ''', (str(year),)).fetchone()
    
    conn.close()
    
    return jsonify({
        'ingresos': [dict(i) for i in ingresos],
        'ingresos_mensuales': [dict(i) for i in ingresos_mensuales],
        'gastos': [dict(g) for g in gastos],
        'gastos_generales': gastos_generales['total'] or 0
    })

@app.route('/api/ingresos-detalle/<int:year>')
def ingresos_detalle(year):
    conn = get_db()
    ingresos = conn.execute('''
        SELECT o.fecha, p.nombre as propiedad, o.precio, o.origen, o.notas,
               strftime('%m', o.fecha) as mes
        FROM ocupaciones o
        JOIN propiedades p ON o.propiedad_id = p.id
        WHERE strftime('%Y', o.fecha) = ?
        ORDER BY o.fecha DESC
    ''', (str(year),)).fetchall()
    conn.close()
    return jsonify([dict(i) for i in ingresos])

@app.route('/api/gastos-detalle/<int:year>')
def gastos_detalle(year):
    conn = get_db()
    gastos = conn.execute('''
        SELECT g.fecha, COALESCE(p.nombre, 'General') as propiedad, 
               g.categoria, g.monto, g.descripcion,
               strftime('%m', g.fecha) as mes
        FROM gastos g
        LEFT JOIN propiedades p ON g.propiedad_id = p.id
        WHERE strftime('%Y', g.fecha) = ?
        ORDER BY g.fecha DESC
    ''', (str(year),)).fetchall()
    conn.close()
    return jsonify([dict(g) for g in gastos])

@app.route('/api/exportar/excel')
def exportar_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Obtener parámetros de filtro
        desde = request.args.get('desde', f'{datetime.now().year}-01-01')
        hasta = request.args.get('hasta', f'{datetime.now().year}-12-31')
        propiedad = request.args.get('propiedad', '')
        
        conn = get_db()
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hoja de Ingresos
        ws1 = wb.active
        ws1.title = "Ingresos"
        headers = ['Fecha', 'Propiedad', 'Precio USD', 'Origen', 'Inquilino']
        ws1.append(headers)
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        query = '''
            SELECT o.fecha, p.nombre, o.precio, o.origen, o.notas
            FROM ocupaciones o
            JOIN propiedades p ON o.propiedad_id = p.id
            WHERE o.fecha >= ? AND o.fecha <= ?
        '''
        params = [desde, hasta]
        
        if propiedad:
            query += ' AND p.nombre = ?'
            params.append(propiedad)
        
        query += ' ORDER BY o.fecha, p.nombre'
        ocupaciones = conn.execute(query, params).fetchall()
        
        total_ingresos = 0
        for o in ocupaciones:
            ws1.append(list(o))
            total_ingresos += o[2] if o[2] else 0
        
        # Fila de total
        ws1.append(['', '', '', '', ''])
        ws1.append(['TOTAL', '', total_ingresos, '', ''])
        
        # Ajustar anchos
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 25
        
        # Hoja de Gastos
        ws2 = wb.create_sheet("Gastos")
        headers = ['Fecha', 'Propiedad', 'Categoría', 'Monto USD', 'Descripción']
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        query_gastos = '''
            SELECT g.fecha, COALESCE(p.nombre, 'General'), g.categoria, g.monto, g.descripcion
            FROM gastos g
            LEFT JOIN propiedades p ON g.propiedad_id = p.id
            WHERE g.fecha >= ? AND g.fecha <= ?
        '''
        params_gastos = [desde, hasta]
        
        if propiedad:
            query_gastos += ' AND (p.nombre = ? OR g.propiedad_id IS NULL)'
            params_gastos.append(propiedad)
        
        query_gastos += ' ORDER BY g.fecha'
        gastos = conn.execute(query_gastos, params_gastos).fetchall()
        
        total_gastos = 0
        for g in gastos:
            ws2.append(list(g))
            total_gastos += g[3] if g[3] else 0
        
        ws2.append(['', '', '', '', ''])
        ws2.append(['TOTAL', '', '', total_gastos, ''])
        
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 30
        
        # Hoja de Resumen
        ws3 = wb.create_sheet("Resumen")
        ws3.append([f'Período: {desde} al {hasta}'])
        ws3.append([f'Propiedad: {propiedad if propiedad else "Todas"}'])
        ws3.append([''])
        
        headers = ['Propiedad', 'Ingresos', 'Noches', 'Ticket Prom', 'Gastos', 'Rentabilidad']
        ws3.append(headers)
        for cell in ws3[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        query_resumen = '''
            SELECT p.nombre,
                   COALESCE(SUM(o.precio), 0) as ingresos,
                   COUNT(o.id) as noches
            FROM propiedades p
            LEFT JOIN ocupaciones o ON p.id = o.propiedad_id 
                AND o.fecha >= ? AND o.fecha <= ?
        '''
        params_res = [desde, hasta]
        
        if propiedad:
            query_resumen += ' WHERE p.nombre = ?'
            params_res.append(propiedad)
        
        query_resumen += ' GROUP BY p.id'
        resumen = conn.execute(query_resumen, params_res).fetchall()
        
        for r in resumen:
            nombre, ingresos, noches = r
            ticket = ingresos / noches if noches > 0 else 0
            gasto_query = '''
                SELECT COALESCE(SUM(monto), 0) FROM gastos 
                WHERE propiedad_id = (SELECT id FROM propiedades WHERE nombre = ?) 
                AND fecha >= ? AND fecha <= ?
            '''
            gasto = conn.execute(gasto_query, (nombre, desde, hasta)).fetchone()[0]
            rentabilidad = ingresos - gasto
            ws3.append([nombre, ingresos, noches, round(ticket, 2), gasto, rentabilidad])
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws3.column_dimensions[col].width = 14
        
        conn.close()
        
        filename = f'data/Reporte_Miami_{desde}_a_{hasta}.xlsx'
        wb.save(filename)
        
        return send_file(filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/descargar-template')
def descargar_template():
    return send_file('data/Template_Alquileres.xlsx', as_attachment=True)

@app.route('/api/importar-excel', methods=['POST'])
def importar_excel():
    try:
        from openpyxl import load_workbook
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Archivo vacío'}), 400
        
        # Leer el Excel
        wb = load_workbook(file)
        ws = wb.active
        
        conn = get_db()
        importados = 0
        errores = []
        origen = request.form.get('origen', 'Dueño')
        
        # Obtener mapeo de propiedades
        props = conn.execute('SELECT id, nombre FROM propiedades').fetchall()
        prop_map = {p['nombre']: p['id'] for p in props}
        
        # Leer filas (empezando desde la 5, saltando headers)
        for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            propiedad, fecha, precio, inquilino = row[0], row[1], row[2], row[3]
            
            # Saltar filas vacías
            if not propiedad or not fecha:
                continue
            
            # Validar propiedad
            if propiedad not in prop_map:
                errores.append(f'Fila {row_num}: Propiedad "{propiedad}" no existe')
                continue
            
            # Convertir fecha si es necesario
            if hasattr(fecha, 'strftime'):
                fecha_str = fecha.strftime('%Y-%m-%d')
            else:
                fecha_str = str(fecha)
            
            # Insertar o actualizar
            try:
                conn.execute('''
                    INSERT OR REPLACE INTO ocupaciones (propiedad_id, fecha, precio, origen, notas)
                    VALUES (?, ?, ?, ?, ?)
                ''', (prop_map[propiedad], fecha_str, float(precio or 0), origen, inquilino or ''))
                importados += 1
            except Exception as e:
                errores.append(f'Fila {row_num}: {str(e)}')
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'importados': importados,
            'errores': errores
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/presentacion/<int:year>')
def generar_presentacion(year):
    conn = get_db()
    
    # Obtener datos
    props = conn.execute('SELECT * FROM propiedades WHERE activo = 1').fetchall()
    
    # Ingresos por propiedad
    ingresos_data = {}
    for p in props:
        data = conn.execute('''
            SELECT COALESCE(SUM(precio), 0) as total,
                   COUNT(*) as noches,
                   origen
            FROM ocupaciones
            WHERE propiedad_id = ? AND strftime('%Y', fecha) = ?
            GROUP BY origen
        ''', (p['id'], str(year))).fetchall()
        
        ingresos_data[p['nombre']] = {
            'total': sum(d['total'] for d in data),
            'noches': sum(d['noches'] for d in data),
            'por_origen': {d['origen']: d['total'] for d in data}
        }
    
    # Gastos por propiedad
    gastos_data = {}
    for p in props:
        total = conn.execute('''
            SELECT COALESCE(SUM(monto), 0) as total
            FROM gastos
            WHERE propiedad_id = ? AND strftime('%Y', fecha) = ?
        ''', (p['id'], str(year))).fetchone()['total']
        gastos_data[p['nombre']] = total
    
    # Gastos generales
    gastos_generales = conn.execute('''
        SELECT COALESCE(SUM(monto), 0) as total
        FROM gastos
        WHERE propiedad_id IS NULL AND strftime('%Y', fecha) = ?
    ''', (str(year),)).fetchone()['total']
    
    conn.close()
    
    # Calcular totales
    total_ingresos = sum(d['total'] for d in ingresos_data.values())
    total_noches = sum(d['noches'] for d in ingresos_data.values())
    total_gastos = sum(gastos_data.values()) + gastos_generales
    total_rentabilidad = total_ingresos - total_gastos
    
    # Colores
    colores = {
        'TIDES 14 B': '#3498db', 'TIDES 5 L': '#e74c3c', 'TIDES 10 L': '#2ecc71',
        'TIDES 10 F': '#9b59b6', 'TIDES 12 F': '#f39c12',
        'Brickell': '#1abc9c', 'Local 1': '#e67e22', 'Local 2': '#34495e'
    }
    
    # Generar HTML
    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Presentación Miami {year}</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap" rel="stylesheet">
    <style>
        *{{margin:0;padding:0;box-sizing:border-box}}
        body{{font-family:'Poppins',sans-serif;background:linear-gradient(135deg,#1a1a2e,#16213e,#0f3460);color:#fff;min-height:100vh}}
        .slide{{min-height:100vh;padding:60px 80px;page-break-after:always}}
        .portada{{display:flex;flex-direction:column;justify-content:center;align-items:center;text-align:center}}
        .portada h1{{font-size:3.5rem;font-weight:700;background:linear-gradient(90deg,#00d2ff,#3a7bd5);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:20px}}
        .portada h2{{font-size:1.8rem;font-weight:300;color:#8892b0;margin-bottom:40px}}
        .portada .loc{{font-size:1.3rem;color:#64ffda}}
        .slide-title{{font-size:2.2rem;font-weight:600;margin-bottom:40px;color:#64ffda;border-left:4px solid #64ffda;padding-left:20px}}
        .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:25px;margin-bottom:40px}}
        .kpi{{background:rgba(255,255,255,0.05);border-radius:16px;padding:25px;text-align:center;border:1px solid rgba(255,255,255,0.1)}}
        .kpi-val{{font-size:2rem;font-weight:700;color:#64ffda}}
        .kpi-lab{{font-size:0.85rem;color:#8892b0;text-transform:uppercase;margin-top:5px}}
        .kpi.warn .kpi-val{{color:#e74c3c}}
        .tabla{{background:rgba(255,255,255,0.03);border-radius:16px;padding:20px;margin-bottom:30px}}
        table{{width:100%;border-collapse:collapse}}
        th{{background:rgba(100,255,218,0.1);padding:12px 15px;text-align:left;color:#64ffda;font-size:0.85rem}}
        td{{padding:12px 15px;border-bottom:1px solid rgba(255,255,255,0.05)}}
        .pos{{color:#2ecc71}}.neg{{color:#e74c3c}}
        .charts{{display:grid;grid-template-columns:1fr 1fr;gap:25px;margin-top:30px}}
        .chart-box{{background:rgba(255,255,255,0.03);border-radius:16px;padding:20px}}
        .chart-title{{font-size:1rem;margin-bottom:15px;color:#8892b0}}
        .footer{{text-align:center;padding:20px;color:#8892b0;font-size:0.9rem}}
        @media print{{.slide{{page-break-after:always}}}}
    </style>
</head>
<body>

<div class="slide portada">
    <h1>Resumen Anual {year}</h1>
    <h2>Propiedades Miami</h2>
    <div class="loc">📍 Miami, Florida</div>
    <p style="margin-top:30px;color:#8892b0">{len([p for p in props if p['tipo']=='departamento'])} Departamentos + {len([p for p in props if p['tipo']=='local'])} Locales</p>
    <p style="margin-top:50px;color:#64ffda;font-size:1.2rem">Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
</div>

<div class="slide">
    <h2 class="slide-title">Resumen Ejecutivo</h2>
    <div class="kpi-grid">
        <div class="kpi"><div class="kpi-val">${total_ingresos:,.0f}</div><div class="kpi-lab">Ingresos Totales</div></div>
        <div class="kpi warn"><div class="kpi-val">${total_gastos:,.0f}</div><div class="kpi-lab">Gastos Totales</div></div>
        <div class="kpi"><div class="kpi-val">${total_rentabilidad:,.0f}</div><div class="kpi-lab">Rentabilidad</div></div>
        <div class="kpi"><div class="kpi-val">{(total_rentabilidad/total_ingresos*100) if total_ingresos > 0 else 0:.1f}%</div><div class="kpi-lab">Margen</div></div>
    </div>
    <div class="kpi-grid">
        <div class="kpi"><div class="kpi-val">{total_noches}</div><div class="kpi-lab">Noches Ocupadas</div></div>
        <div class="kpi"><div class="kpi-val">{(total_noches/(365*len(props))*100):.1f}%</div><div class="kpi-lab">Ocupación Promedio</div></div>
        <div class="kpi"><div class="kpi-val">${total_ingresos/total_noches if total_noches > 0 else 0:.0f}</div><div class="kpi-lab">Ticket Promedio</div></div>
        <div class="kpi"><div class="kpi-val">{len(props)}</div><div class="kpi-lab">Propiedades</div></div>
    </div>
</div>

<div class="slide">
    <h2 class="slide-title">Detalle por Propiedad</h2>
    <div class="tabla">
        <table>
            <tr><th>Propiedad</th><th>Ingresos</th><th>Gastos</th><th>Rentabilidad</th><th>Noches</th><th>% Ocup</th><th>Ticket</th></tr>'''
    
    for p in props:
        nombre = p['nombre']
        ing = ingresos_data.get(nombre, {'total': 0, 'noches': 0})
        gast = gastos_data.get(nombre, 0)
        rent = ing['total'] - gast
        noches = ing['noches']
        ticket = ing['total'] / noches if noches > 0 else 0
        ocup = (noches / 365) * 100
        color = colores.get(nombre, '#666')
        cls = 'pos' if rent > 0 else 'neg'
        
        html += f'''<tr>
            <td style="border-left:4px solid {color};padding-left:15px;font-weight:600">{nombre}</td>
            <td>${ing['total']:,.0f}</td>
            <td>${gast:,.0f}</td>
            <td class="{cls}">${rent:,.0f}</td>
            <td>{noches}</td>
            <td>{ocup:.1f}%</td>
            <td>${ticket:.0f}</td>
        </tr>'''
    
    html += f'''<tr style="background:rgba(100,255,218,0.1);font-weight:600">
            <td>TOTAL</td>
            <td>${total_ingresos:,.0f}</td>
            <td>${total_gastos:,.0f}</td>
            <td class="pos">${total_rentabilidad:,.0f}</td>
            <td>{total_noches}</td>
            <td>{(total_noches/(365*len(props))*100):.1f}%</td>
            <td>${total_ingresos/total_noches if total_noches > 0 else 0:.0f}</td>
        </tr>
        </table>
    </div>
</div>

<div class="slide">
    <h2 class="slide-title">Ingresos por Origen</h2>
    <div class="tabla">
        <table>
            <tr><th>Propiedad</th><th>Dueño</th><th>Alicia</th><th>Estanislao</th><th>Total</th></tr>'''
    
    total_dueño = total_alicia = total_estanislao = 0
    for p in props:
        nombre = p['nombre']
        ing = ingresos_data.get(nombre, {'total': 0, 'por_origen': {}})
        dueño = ing['por_origen'].get('Dueño', 0)
        alicia = ing['por_origen'].get('Alicia', 0)
        estanislao = ing['por_origen'].get('Estanislao', 0)
        total_dueño += dueño
        total_alicia += alicia
        total_estanislao += estanislao
        
        html += f'''<tr>
            <td>{nombre}</td>
            <td>${dueño:,.0f}</td>
            <td>${alicia:,.0f}</td>
            <td>${estanislao:,.0f}</td>
            <td><strong>${ing['total']:,.0f}</strong></td>
        </tr>'''
    
    html += f'''<tr style="background:rgba(100,255,218,0.1);font-weight:600">
            <td>TOTAL</td>
            <td>${total_dueño:,.0f}</td>
            <td>${total_alicia:,.0f}</td>
            <td>${total_estanislao:,.0f}</td>
            <td><strong>${total_ingresos:,.0f}</strong></td>
        </tr>
        </table>
    </div>
    
    <div class="charts">
        <div class="chart-box">
            <div class="chart-title">📊 Distribución por Origen</div>
            <canvas id="chart1"></canvas>
        </div>
        <div class="chart-box">
            <div class="chart-title">📈 Ingresos por Propiedad</div>
            <canvas id="chart2"></canvas>
        </div>
    </div>
</div>

<div class="slide">
    <h2 class="slide-title">Conclusiones</h2>
    <div style="font-size:1.2rem">
        <div style="padding:25px 0;border-bottom:1px solid rgba(255,255,255,0.1)">
            ✅ <strong>Rentabilidad Total:</strong> ${total_rentabilidad:,.0f} ({(total_rentabilidad/total_ingresos*100) if total_ingresos > 0 else 0:.1f}% margen)
        </div>
        <div style="padding:25px 0;border-bottom:1px solid rgba(255,255,255,0.1)">
            📊 <strong>Ingresos por Dueño:</strong> ${total_dueño:,.0f} ({(total_dueño/total_ingresos*100) if total_ingresos > 0 else 0:.1f}%)
        </div>
        <div style="padding:25px 0;border-bottom:1px solid rgba(255,255,255,0.1)">
            👥 <strong>Ingresos por Terceros:</strong> ${total_alicia + total_estanislao:,.0f} ({((total_alicia + total_estanislao)/total_ingresos*100) if total_ingresos > 0 else 0:.1f}%)
        </div>
        <div style="padding:25px 0;border-bottom:1px solid rgba(255,255,255,0.1)">
            🛏️ <strong>Ocupación:</strong> {total_noches} noches ({(total_noches/(365*len(props))*100):.1f}% promedio)
        </div>
        <div style="padding:25px 0">
            💵 <strong>Ticket Promedio:</strong> ${total_ingresos/total_noches if total_noches > 0 else 0:.0f} por noche
        </div>
    </div>
</div>

<div class="footer">
    Generado automáticamente desde Plataforma Alquileres Miami | {datetime.now().strftime('%d/%m/%Y %H:%M')}
</div>

<script>
Chart.defaults.color = '#8892b0';
new Chart(document.getElementById('chart1'), {{
    type: 'doughnut',
    data: {{
        labels: ['Dueño', 'Alicia', 'Estanislao'],
        datasets: [{{
            data: [{total_dueño}, {total_alicia}, {total_estanislao}],
            backgroundColor: ['#2ecc71', '#9b59b6', '#f1c40f']
        }}]
    }}
}});

new Chart(document.getElementById('chart2'), {{
    type: 'bar',
    data: {{
        labels: {json.dumps([p['nombre'] for p in props])},
        datasets: [{{
            data: {json.dumps([ingresos_data.get(p['nombre'], {}).get('total', 0) for p in props])},
            backgroundColor: {json.dumps([colores.get(p['nombre'], '#666') for p in props])}
        }}]
    }},
    options: {{ plugins: {{ legend: {{ display: false }} }} }}
}});
</script>
</body>
</html>'''
    
    return Response(html, mimetype='text/html')

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🏠 PLATAFORMA ALQUILERES MIAMI")
    print("="*50)
    print("\n📍 Abrí tu navegador en: http://localhost:5000")
    print("\n💡 Para cerrar: Ctrl+C")
    print("="*50 + "\n")
    app.run(debug=True, port=5000)
